from openpyxl import load_workbook

wb = load_workbook("data/GrowthPilot_Sales_Dataset.xlsx")
print(wb.sheetnames)